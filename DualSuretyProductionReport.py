# activate venv: cd "C:\Users\ssrsserviceaccount\Documents\Oleg Files\Python\virtual_env"; .venv\Scripts\Activate.ps1  --old
# cd "C:\Users\ssrsserviceaccount\Documents\Oleg Files\Python"; .venv\Scripts\Activate.ps1

import sqlalchemy as sa
import urllib
import shutil
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment 
from functions import write_summary_formula, apply_double_border_range, double_border_format,adjust_columns_width, header_formatting

# Load environment variables from .env
load_dotenv()

# Get credentials from environment
server = os.getenv("SQL_SERVER")
database = os.getenv("SQL_DB")
user = os.getenv("SQL_USER")
password = os.getenv("SQL_PASSWORD")

##\\fs01\Align\Accounting\Accounting\Carrier Reporting\FY2024\DUAL NA\AXIS Insurance (Surety)

# # Get the directory of the current project
project_folder = os.path.abspath(os.curdir)+ '\\' +'Dual Surety Production Report'

params = urllib.parse.quote_plus("Driver={ODBC Driver 17 for SQL Server};"
                                 "SERVER={server};"
                                 "DATABASE={database};"
                                 "UID={user};"
                                 "PWD={password}")

engine = sa.create_engine("mssql+pyodbc:///?odbc_connect={}".format(params))

# read .sql file
#sql01 = open('get_params.sql', 'r') 
sql01 = open(r'get_params.sql', 'r') 
sqlFile = sql01.read()
sql01.close()

# save parameters to df
df_params = pd.read_sql(sqlFile, con=engine)





# convert DateFrom to m-yyyy to use in folder name
# grab DateFrom only for furthure converstion
date_from = df_params.iloc[0,0]
year = str(date_from.year)
month_year = f"{date_from.month:02d}.{date_from.year}"   #str(date_from.month) + '.' + str(date_from.year)


# substitute empty values to NULL for passing in SP
def NoneToNull(s):
    return "NULL" if s is None else "'" + str(s) + "'"

# execute sp for each row of df_params and supply parameter values
for idx in df_params.index:
    query = (
        "SET NOCOUNT ON; EXEC Oleg_MEJames_spFin_rptProductionReportOnDetailForCatalityc "
        + "@DateFrom =" + "'" + str(df_params.loc[idx, "DateFrom"]) + "'" + ","
        + "@DateTo =" + "'" + str(df_params.loc[idx, "DateTo"]) + "'" + ","
        + "@CompanyLocationGuids =" + NoneToNull(df_params.loc[idx, "CompanyLocationGuids"]) + ","
        + "@CarrierLocationGuids =" + NoneToNull(df_params.loc[idx, "CarrierLocationGuids"]) + ","
        + "@LineGuid =" + NoneToNull(df_params.loc[idx, "LineGuid"]) + ","
        + "@OfficeGuid =" + NoneToNull(df_params.loc[idx, "officeGuid"]) + ","
        + "@ShowAllOffices =" + str(df_params.loc[idx, "ShowAllOffices"])
    )

    # save data into dataframe
    df = pd.read_sql(query, con=engine)

    # keep excel file in project folder for formatting
    file_name = f"AXIS Production Report {month_year}.xlsx"
    file_path = f"{project_folder}\\{file_name}"
    df.to_excel(file_path, index=False, startrow=8) # will start from row 9

    # create workbook. Basically saying which file will be modified
    wb = openpyxl.load_workbook(file_path)
   
    # open sheet to write into. 
    ws = wb.active

    ####################### Formatting ########################

    # put excel header
    header_formatting(ws, df=df_params)

    # # Freeze the rows above and the columns to the left of B2
    ws.freeze_panes = "A10" 

    # Iterate through the header row and remove the borders
    for cell in ws[9]:
        cell.border = openpyxl.styles.borders.Border()
        cell.alignment = Alignment(horizontal='left')

    # get first and last row of 
    first_data_row = ws.min_row + 1 # gives first value row after column header
    last_data_row = ws.max_row 

    # change format to Accounting wihtout $ for columns N-AB 
    for col in ws.iter_cols(min_col=14, max_col=28, min_row=first_data_row,  max_row=last_data_row) : # range where to start and end loop
        for cell in col: 
            cell.number_format = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

    # change format to Accounting wihtout $ for columns AE-AF 
    for col in ws.iter_cols(min_col=31, max_col=32, min_row=first_data_row,  max_row=last_data_row) : # range where to start and end loop
        for cell in col: 
            cell.number_format = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'

    write_summary_formula(ws, min_row = last_data_row + 2, 
                            min_col=14, 
                            max_row = last_data_row + 2,
                            max_col = 24,
                            sum_startrow = first_data_row,
                            sum_endrow = last_data_row)

    write_summary_formula(ws, min_row = last_data_row + 2, 
                            min_col=26, 
                            max_row = last_data_row + 2,
                            max_col = 32,
                            sum_startrow = first_data_row,
                            sum_endrow = last_data_row)


    # double border starting from column "Peril" to "Net Due Carrier"
    apply_double_border_range(ws, start_row =last_data_row +2, 
                                start_column = 14,
                                end_row =last_data_row +2,
                                end_column = 32,
                                double_border=double_border_format)

    adjust_columns_width(ws, ['B'], 25)
    adjust_columns_width(ws, ['D', 'E', 'F', 'G'], 13)
    adjust_columns_width(ws, ['I', 'J'], 14)
    adjust_columns_width(ws, ['N', 'O', 'P', 'Q', 'R','S', 'T', 'U', 'V', 'W','X', 'Y', 'Z', 'AA', 'AB', 'AE', 'AF'], 13)


    #replace_zeros_in_columns(ws, ['AE', 'AF'] )


    # save changes
    wb.save(file_path)

    # extract destination path
    dest_path = df_params.loc[idx, 'path']

# uncomment below when production ready
''' 
    # if destination path does not exists then create it and move file, oterwise just move file
    if not os.path.exists(dest_path): 
        os.mkdir(dest_path)
        shutil.move(file_name, dest_path + '\\' + file_name)     
    else:
        shutil.move(file_name, dest_path + '\\' + file_name)
'''
