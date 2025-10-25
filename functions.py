import pandas as pd
import openpyxl
from openpyxl.styles import  Font, Border, Side, Alignment, numbers 
from openpyxl import load_workbook,Workbook
from openpyxl.worksheet.worksheet import Worksheet

def write_summary_formula(ws, min_row, min_col, max_row, max_col,sum_startrow,sum_endrow):
# function to dynamically write sum formula
# min_row: row where sum should be placed
# min_col: column where sum should be placed
# sum_startrow: row from where sum calculation should be starting
# sum_endrow: row from where sum calculation should be ending
    for col in ws.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col): # range where to start and end loop
        for cell in col:
            cell_sum_start = cell.column_letter + str(sum_startrow)
            cell_sum_end = cell.column_letter + str(sum_endrow)
            cell.value = f'=SUM({cell_sum_start}:{cell_sum_end})'
            cell.number_format = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)' # Acconting stype without $. Source: https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
            cell.font = Font(bold=True)
            #cell.border = double_border


# create double border for summary row
def apply_double_border_range(ws, start_row, start_column, end_row, end_column, double_border):
    # Iterate over the range of cells and apply double border
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            ws.cell(row=row, column=column).border = double_border

# top and double buttom border style
double_border_format = Border(top=Side(style='thin'), bottom=Side(style='double'))


# adjust column widts. Pass multiple columns Ex: ['A', 'B', 'C']
def adjust_columns_width(ws, columns_to_adjust, width):
    for col in columns_to_adjust:
        ws.column_dimensions[col].width = width

from openpyxl.styles import numbers
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------
# format number to percentage
# ---------------------------------------------------------------------

def format_column_as_percentage(
    ws: Worksheet, 
    col_letter_or_index, 
    first_row: int, 
    last_row: int, 
    decimals: int = 2
):
    """
    Format a column in an Excel worksheet as percentage.

    Args:
        ws: openpyxl Worksheet object
        col_letter_or_index: Column letter (e.g., 'J') or 1-based index (e.g., 10)
        first_row: First row to apply formatting
        last_row: Last row to apply formatting
        decimals: Number of decimals for percentage formatting
    """
    # Determine number format
    fmt = numbers.FORMAT_PERCENTAGE_00 if decimals == 2 else numbers.FORMAT_PERCENTAGE

    # Convert letter to column index if needed
    if isinstance(col_letter_or_index, str):
        from openpyxl.utils import column_index_from_string
        col_index = column_index_from_string(col_letter_or_index)
    else:
        col_index = col_letter_or_index

    # Apply formatting
    for col in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=first_row, max_row=last_row):
        for cell in col:
            if cell.value is not None:
                cell.value /= 100
                cell.number_format = fmt

# # Example: format column J (10) from row 7 to 50 as percentage
# format_column_as_percentage(ws, 'J', first_row=7, last_row=50)

# ---------------------------------------------------------------------
# write header
# ---------------------------------------------------------------------

# def header_formatting(ws, df):
#     # write Carrier name 
#     carrier = ws.cell(row=1, column=1)
#     carrier.value = 'AXIS Insurance Company'
#     carrier.font = Font(name='Arial', bold=True, size = "10")

#     # write office name
#     office = ws.cell(row=2, column=1)
#     office.value = 'DUAL North America'
#     office.font = Font(name='Arial',bold=True, size = "10")

#     # write program name
#     program = ws.cell(row=3, column=1)
#     program.value = 'Surety Bond Program'
#     program.font = Font(name='Arial',bold=True, size = "10")

#     # write prod report
#     prodreport = ws.cell(row=4, column=1)
#     prodreport.value = 'Production Report'
#     prodreport.font = Font(name='Arial', bold=True, size = "10")

#     # grab top 1 DateFrom and DateTo 
#     dateFrom = df['DateFrom'].iloc[0]
#     dateTo = df['DateTo'].iloc[0]
#     # Format dates in mm/dd/yyyy format
#     dateFrom_str = dateFrom.strftime('%m/%d/%Y')
#     dateTo_str = dateTo.strftime('%m/%d/%Y')
#     dateFromTo = f"For the period {dateFrom_str} to {dateTo_str}"
#     ws['A5'].value = dateFromTo
#     ws['A5'].font = Font(bold=True)


#     # since iterating through each row of df individually - its ok to grab top 1 "PaymentDue"
#     paymentDue = df['PaymentDue'].iloc[0]
#     paymentTermsDue_str = paymentDue.strftime('%m/%d/%Y')
#     paymentTermsDue = f"Payment Terms Due: {paymentTermsDue_str}"
#     ws['A6'].value = paymentTermsDue
#     ws['A6'].font = Font(bold=True)

# ---------------------------------------------------------------------
# Writes a piece of text into a given (row, column) on the worksheet ws, and makes it bold.
# ---------------------------------------------------------------------
def write_bold(ws, row, col, text, size=10):
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(name='Arial', bold=True, size=size)


# ---------------------------------------------------------------------
# "Account Current" header
# ---------------------------------------------------------------------

def header_formatting(ws, row):
    """Write dynamic header info from the current parameter row."""
    headers = [
        row['OfficeName'],
        row['CarrierLocationNames'],
        "Account Current"
    ]
    print(headers)
    for i, text in enumerate(headers, start=1):
        write_bold(ws, i, 1, text)

    # Format date range
    dateFrom = row['DateFrom'].strftime('%m/%d/%Y')
    dateTo = row['DateTo'].strftime('%m/%d/%Y')
    write_bold(ws, 4, 1, f"For the period {dateFrom} to {dateTo}")

    # Payment due
    paymentDue = row['PaymentDue'].strftime('%m/%d/%Y')
    write_bold(ws, 5, 1, f"Payment Terms Due: {paymentDue}")


# ---------------------------------------------------------------------
# define Account Current header
# ---------------------------------------------------------------------

#def account_current_header_formatting(ws, df):

    # # write office name
    # office = ws.cell(row=2, column=1)
    # office.value = 'DUAL North America'
    # office.font = Font(name='Arial',bold=True, size = "10")

    # # write Carrier name 
    # carrier = ws.cell(row=1, column=1)
    # #carrier.value = 'AXIS Insurance Company'
    # carrier.value = df['CarrierLocationNames'].iloc[0]
    # carrier.font = Font(name='Arial', bold=True, size = "10")

    # # write "Account Current" in a cell 3
    # AccountCurrent = ws.cell(row=2, column=1)
    # AccountCurrent.value = 'Account Current'
    # AccountCurrent.font = Font(bold=True)

    # # write program name
    # program = ws.cell(row=3, column=1)
    # program.value = 'Surety Bond Program'
    # program.font = Font(name='Arial',bold=True, size = "10")

    # # write prod report
    # prodreport = ws.cell(row=4, column=1)
    # prodreport.value = 'Production Report'
    # prodreport.font = Font(name='Arial', bold=True, size = "10")

    # # grab top 1 DateFrom and DateTo 
    # dateFrom = df['DateFrom'].iloc[0]
    # dateTo = df['DateTo'].iloc[0]

    # # Format dates in mm/dd/yyyy format
    # dateFrom_str = dateFrom.strftime('%m/%d/%Y')
    # dateTo_str = dateTo.strftime('%m/%d/%Y')
    # dateFromTo = f"For the period {dateFrom_str} to {dateTo_str}"
    # ws['A5'].value = dateFromTo
    # ws['A5'].font = Font(bold=True)


    # # since iterating through each row of df individually - its ok to grab top 1 "PaymentDue"
    # paymentDue = df['PaymentDue'].iloc[0]
    # paymentTermsDue_str = paymentDue.strftime('%m/%d/%Y')
    # paymentTermsDue = f"Payment Terms Due: {paymentTermsDue_str}"
    # ws['A6'].value = paymentTermsDue
    # ws['A6'].font = Font(bold=True)



def replace_zeros_in_columns(ws, columns_to_replace):
    """
    Replace 0 with '-' in specified columns of a given worksheet.

    Parameters:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet object to modify.
    columns_to_replace (list): List of column letters to replace 0 with '-'.
    """
    # Iterate through the rows in the sheet
    for row in ws.iter_rows():
        for col_letter in columns_to_replace:
            cell = row[openpyxl.utils.cell.column_index_from_string(col_letter) - 1]
            if cell.value == 0:
                cell.value = '-'
                cell.alignment = Alignment(horizontal='right')


