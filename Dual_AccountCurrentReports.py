"""
Dual Surety Production Report Generator
---------------------------------------
1. Reads SQL parameters from `get_params.sql`
2. Executes stored procedure for each parameter set
3. Exports and formats Excel reports
4. Optionally moves files to final destination

Author: Oleg
"""
#  Activate .venv:  cd "C:\Users\ssrsserviceaccount\Documents\Oleg Files\Python"; .venv\Scripts\Activate.ps1


import os
import shutil
import logging
import urllib
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv

import pandas as pd
import sqlalchemy as sa
import openpyxl
from openpyxl.styles import Alignment, Font

from functions import (
    write_summary_formula,
    apply_double_border_range,
    double_border_format,
    adjust_columns_width,
  #  header_formatting,
    format_column_as_percentage
)

# ---------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------
load_dotenv()  # Loads environment variables from .env file if present

# Configure logging
logging.basicConfig(
    filename="dual_reports.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Project directory
PROJECT_DIR = Path.cwd()
SQL_FILE = PROJECT_DIR / "get_params.sql"


# ---------------------------------------------------------------------
# Database Connection
# ---------------------------------------------------------------------
def get_engine() -> sa.engine.Engine:
    """Create a SQLAlchemy engine using secure credentials."""
    server = os.getenv("SQL_SERVER")
    database = os.getenv("SQL_DB")
    user = os.getenv("SQL_USER")
    password = os.getenv("SQL_PASSWORD")

    if not password:
        raise ValueError("❌ Missing SQL_PASSWORD environment variable")

    params = urllib.parse.quote_plus(
        f"Driver={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};DATABASE={database};UID={user};PWD={password}"
    )
    print("SQL DB connection succesful")
    return sa.create_engine(f"mssql+pyodbc:///?odbc_connect={params}")
    

# ---------------------------------------------------------------------
# SQL / Parameter Handling
# ---------------------------------------------------------------------
def get_params(engine: sa.engine.Engine) -> pd.DataFrame:
    """Load parameters for stored procedure from .sql file."""
    with open(SQL_FILE, "r") as f:
        sql_query = f.read()
    return pd.read_sql(sql_query, con=engine)


def none_to_null(value):
    """Convert Python None/NaN to SQL NULL string."""
    if pd.isna(value) or value is None:
        return "NULL"
    return f"'{value}'"


def build_query(row: pd.Series) -> str:
    """Build SQL EXEC statement for stored procedure."""
    return (
        "SET NOCOUNT ON; EXEC Oleg_MEJames_spFin_rptProductionReportOnDetailForCatalityc "
        f"@DateFrom='{row.DateFrom}', "
        f"@DateTo='{row.DateTo}', "
        f"@CompanyLocationGuids={none_to_null(row.CompanyLocationGuids)}, "
        f"@CarrierLocationGuids={none_to_null(row.CarrierLocationGuids)}, "
        f"@LineGuid={none_to_null(row.LineGuid)}, "
        f"@OfficeGuid={none_to_null(row.officeGuid)}, "
        f"@ShowAllOffices={row.ShowAllOffices}"
    )
# ---------------------------------------------------------------------
# Excel Formatting
# ---------------------------------------------------------------------
def format_currency_columns(ws, start_col, end_col, first_row, last_row):
    """Apply accounting format (no $) to a column range."""
    num_format = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'
    for col in ws.iter_cols(min_col=start_col, max_col=end_col,
                            min_row=first_row, max_row=last_row):
        for cell in col:
            cell.number_format = num_format

def write_bold(ws, row, col, text, size=10):
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(name='Arial', bold=True, size=size)


def header_formatting(ws, row):
    """Write dynamic header info from the current parameter row."""
    headers = [
        row['OfficeName'],
        row['CarrierLocationNames'],
        "Account Current"
    ]
    print(headers)
    for i, text in enumerate(headers, start=1):
        write_bold(ws, i, 1, text)

    # Format date range
    dateFrom = row['DateFrom'].strftime('%m/%d/%Y')
    dateTo = row['DateTo'].strftime('%m/%d/%Y')
    write_bold(ws, 4, 1, f"For the period {dateFrom} to {dateTo}")

    # Payment due
    paymentDue = row['PaymentDue'].strftime('%m/%d/%Y')
    write_bold(ws, 5, 1, f"Payment Terms Due: {paymentDue}")


def format_excel(ws, row: pd.Series):
    """Apply header, borders, alignment, and summary formulas."""
    #header_formatting(ws, row)
    # Layout settings
    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False

    # Style header row 7 (column titles)
    for cell in ws[7]:
        cell.border = openpyxl.styles.borders.Border()
        cell.alignment = Alignment(horizontal='left', wrapText=True)
    ws.row_dimensions[7].height = 30

    # Format data section
    first_data_row = ws.min_row + 1 # gives 8
    last_data_row = ws.max_row
    format_currency_columns(ws, 11, 13, first_data_row, last_data_row)
    format_column_as_percentage(ws, 'J', first_row=first_data_row, last_row=last_data_row)

    # Summary and totals
    write_summary_formula(ws, min_row=last_data_row + 2,
                          min_col=11, max_row=last_data_row + 2,
                          max_col=13, sum_startrow=first_data_row, sum_endrow=last_data_row)

    # Borders and column widths
    apply_double_border_range(ws, start_row=last_data_row + 2, start_column=11,
                              end_row=last_data_row + 2, end_column=13,
                              double_border=double_border_format)

    ws.cell(row=last_data_row+2, column=10, value="Total").font = Font(bold=True)

    # Adjust widths
    adjust_columns_width(ws, ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M'], 15)

    # # Header
    # headers = [
    #     row['OfficeName'],
    #     row['CarrierLocationNames'],
    #     "Account Current"
    # ]
    # print(headers)
    # for i, text in enumerate(headers, start=1):
    #     write_bold(ws, i, 1, text)
    
    
    # dateFrom = row['DateFrom'].strftime('%m/%d/%Y')
    # dateTo = row['DateTo'].strftime('%m/%d/%Y')
    # write_bold(ws, 4, 1, f"For the period {dateFrom} to {dateTo}")

    # # Payment due
    # paymentDue = row['PaymentDue'].strftime('%m/%d/%Y')
    # write_bold(ws, 5, 1, f"Payment Terms Due: {paymentDue}")

    # Call header formatting with the current row
    header_formatting(ws, row)
 


# ---------------------------------------------------------------------
# File Handling
# ---------------------------------------------------------------------
def save_excel(df: pd.DataFrame, row: pd.Series, file_name: str, dest_path: str):
    """Save DataFrame to Excel, format it, and move to destination."""
    file_path = PROJECT_DIR / file_name

    # Write DataFrame to Excel starting from 7th row (headers above)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=6) 

    # Apply formatting
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    #header_formatting(ws, row)
    format_excel(ws, row)  # pass current row instead of full df_params
    wb.save(file_path)

    # Move to destination
    dest_dir = Path(dest_path)
    dest_dir.mkdir(parents=True, exist_ok=True)
    shutil.move(str(file_path), dest_dir / file_name)


# ---------------------------------------------------------------------
# rename columns from raw (sql) to be displayed in Excel
# ---------------------------------------------------------------------

columns_mapping  = {'MultiCarrierContractNumber': 'MultiCarrierContractNumber'
                    ,'PolicyNumber': 'Policy Number'
                    ,'Insured': 'Named Insured'
                    ,'StateID': 'State'
                    ,'EffectiveDate': 'Effective'
                    ,'ExpirationDate': 'Expiration'
                    ,'InvoiceDate': 'Bill Date'
                    ,'PolicyType': 'Type'
                    ,'LineOfCoverage': 'Peril'
                    ,'Commission': 'Company Commission'
                    ,'Premium': 'Gross Premium'
                    ,'TotalComission': 'Gross Commission'
                    ,'NetDueCarrier': 'Net Due Carrier'
                    }

# Keep only the needed columns (ignore missing ones gracefully)
selected_columns = [col for col in columns_mapping.keys()]

# ---------------------------------------------------------------------
# Main Logic
# ---------------------------------------------------------------------
def main():
    try:
        engine = get_engine()
        df_params = get_params(engine)
        logging.info(f"Loaded {len(df_params)} parameter sets from SQL file")

        with engine.begin() as conn:
            for idx, row in df_params.iterrows():
                file_name = row["FileName"]
                dest_path = row["path"]

                logging.info(f"Processing {idx+1}/{len(df_params)}: {file_name}")
                query = build_query(row)

                try:
                    df = pd.read_sql(query, con=conn)
                    # Keep only selected columns that exist in df
                    df = df[[col for col in selected_columns if col in df.columns]]
                    # Rename columns using your mapping
                    df = df.rename(columns={col: columns_mapping[col] for col in df.columns})
                    # Pass the current row (parameters) to save_excel
                    save_excel(df, row, file_name, dest_path)  # or PROJECT_DIR
                    logging.info(f"Successfully generated {file_name}")
                except Exception as e:
                    logging.error(f"Failed to process {file_name}: {e}")

        logging.info("All reports processed successfully.")
    except Exception as e:
        logging.critical(f"Fatal error: {e}")



if __name__ == "__main__":
    main()

